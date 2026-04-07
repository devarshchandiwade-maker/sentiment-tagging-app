import streamlit as st
import pandas as pd
import google.generativeai as genai
import time
from google.api_core.exceptions import ResourceExhausted
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ----------------------------
# Gemini API Key
# ----------------------------
genai.configure(api_key="AIzaSyDNCHihIix6RuUW0hMxObR97NFWv6JV1I8")

model = genai.GenerativeModel("gemini-2.5-flash")

# ----------------------------
# Streamlit UI
# ----------------------------

st.set_page_config(
    page_title="Gozoop AI Sentiment Analytics",
    page_icon="https://media.licdn.com/dms/image/v2/C510BAQGIE1KrFwlsZA/company-logo_200_200/company-logo_200_200/0/1631433137278/gozoop_pvt_ltd_logo?e=2147483647&v=beta&t=f5MM3KJNijSFtWzJMXQmcJqVp_psiLhT2pPxLDTOsss",
    layout="centered"
)

st.image(
    "https://media.licdn.com/dms/image/v2/C510BAQGIE1KrFwlsZA/company-logo_200_200/company-logo_200_200/0/1631433137278/gozoop_pvt_ltd_logo?e=2147483647&v=beta&t=f5MM3KJNijSFtWzJMXQmcJqVp_psiLhT2pPxLDTOsss",
    width=200
)

st.title("AI Sentiment & Tagging Analytics")



st.write("Upload Excel/CSV with **Comment** and **Source** columns")
st.write("Only **1 excel sheet with 10 rows are process othervise give error.** ")

uploaded_file = st.file_uploader("Upload File", type=["xlsx", "csv"])

# ----------------------------
# Gemini Function
# ----------------------------
def analyze_comment(comment, tag_list):

    tag_string = ", ".join(tag_list)

    prompt = f"""
        Analyze the following social media comment.

        Comment: {comment}

        Existing Tags:
        {tag_string}

        Tasks:

        1. Generate a short and meaningful tag.
        2. Detect sentiment: Positive, Negative, or Neutral.
        3. Detect abusive or offensive language.

        Abuse Detection Rules:

        - Abuse means insulting, targeting, or attacking a person, brand, show, or group.
        - Abuse must be directed at someone or something.
        - If offensive or vulgar words are used in excitement, hype, appreciation, or positive expression, do not mark as abuse.
        - Abuse should only be marked when the intention is insulting, offensive, or harmful.
        - If abuse is present → Abuse = Yes and Sentiment = Negative.
        - If vulgar/slang words are used positively or emotionally → Abuse = No.

        Tag Rules:

        - Keep tag short and clear.
        - Avoid generic tags.
        - Reuse existing tags if meaning matches.
        - If abuse present → Tag should reflect abuse or complaint.

        Sentiment Rules:

        Positive → excitement, appreciation, support, hype, love  
        Negative → complaint, anger, insult, offensive language  
        Neutral → question, suggestion, general statement  

        Return strictly in this format:

        Tag: short sentence
        Sentiment: Positive/Negative/Neutral
        Abuse: Yes/No
        """
    try:
        response = model.generate_content(prompt)
        text = response.text

    except ResourceExhausted:
        st.warning("Quota exceeded. Waiting 60 seconds...")
        time.sleep(60)
        response = model.generate_content(prompt)
        text = response.text

    tag = ""
    sentiment = ""
    abuse = "No"

    for line in text.split("\n"):

        if "Tag:" in line:
            tag = line.replace("Tag:", "").strip()

        if "Sentiment:" in line:
            sentiment = line.replace("Sentiment:", "").strip()

        if "Abuse:" in line:
            abuse = line.replace("Abuse:", "").strip()

    return tag, sentiment, abuse


# ----------------------------
# Auto Processing
# ----------------------------
if uploaded_file:

    # Read file
    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    st.subheader("Preview Data")
    st.dataframe(df.head())

    if "Comment" not in df.columns:
        st.error("Comment column not found")
        st.stop()

    # ----------------------------
    # Session State
    # ----------------------------
    if "processed_rows" not in st.session_state:
        st.session_state.processed_rows = 0

    if "result_df" not in st.session_state:
        st.session_state.result_df = pd.DataFrame()

    if "tag_list" not in st.session_state:
        st.session_state.tag_list = []

    # ----------------------------
    # Auto Processing
    # ----------------------------
    if st.button("Start Processing"):

        progress_bar = st.progress(0)
        status_text = st.empty()

        total_rows = len(df)
        batch_size = 5

        while st.session_state.processed_rows < total_rows:

            start = st.session_state.processed_rows
            end = start + batch_size

            batch_df = df.iloc[start:end]

            status_text.info(
                f"Processing rows {start+1} to {min(end, total_rows)}"
            )

            tags_list = []
            sentiment_list = []
            abuse_list = []

            for i, row in batch_df.iterrows():

                comment = row["Comment"]

                tag, sentiment, abuse = analyze_comment(
                    comment,
                    st.session_state.tag_list
                )

                if tag and tag not in st.session_state.tag_list:
                    st.session_state.tag_list.append(tag)

                tags_list.append(tag)
                sentiment_list.append(sentiment)
                abuse_list.append(abuse)

                time.sleep(10)

            batch_df["Tag"] = tags_list
            batch_df["Sentiment"] = sentiment_list
            batch_df["Abuse"] = abuse_list

            st.session_state.result_df = pd.concat(
                [st.session_state.result_df, batch_df],
                ignore_index=True
            )

            st.session_state.processed_rows += batch_size

            progress = st.session_state.processed_rows / total_rows
            progress_bar.progress(min(progress, 1.0))

        status_text.success("All rows processed successfully")

            # ----------------------------
        # Show Results
        # ----------------------------
        st.subheader("Processed Data")
        st.dataframe(st.session_state.result_df)

        output_file = "sentiment_result.xlsx"
        st.session_state.result_df.to_excel(output_file, index=False)

        # ----------------------------
        # Apply Red Color to Abuse Rows
        # ----------------------------
        wb = load_workbook(output_file)
        ws = wb.active

        red_fill = PatternFill(
            start_color="FFCCCC",
            end_color="FFCCCC",
            fill_type="solid"
        )

        for row in range(2, ws.max_row + 1):

            abuse_value = ws[f"D{row}"].value

            if abuse_value == "Yes":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = red_fill

        wb.save(output_file)

        # ----------------------------
        # Download Button
        # ----------------------------
        with open(output_file, "rb") as f:
            st.download_button(
                label="Download Result Excel",
                data=f,
                file_name="sentiment_result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # ----------------------------
        # Tag List
        # ----------------------------
        if st.session_state.tag_list:

            st.subheader("Generated Tag List")

            tag_df = pd.DataFrame(
                st.session_state.tag_list,
                columns=["Tags"]
            )

            st.dataframe(tag_df)
